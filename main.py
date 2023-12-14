import prettytable as pt
import os
from prettytable.colortable import ColorTable, Themes
import csv
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import PatternFill


tb1 = pt.PrettyTable()
tb1.align = "l"   # 靠左對齊


# 選擇風格
tb1.set_style(pt.SINGLE_BORDER)
# tb1.set_style(pt.DOUBLE_BORDER)
# tb1.set_style(pt.PLAIN_COLUMNS)
# tb1.set_style(pt.MSWORD_FRIENDLY)


# 定義欄位格式&擷取TXT檔特定內容
tb1.field_names = ["Marketing Name", "S/N", "System BIOS", "ME Firmware", "Feature Byte"]
folder_path = "./BCU_Files"
for filename in os.listdir(folder_path):
    if filename.endswith(".txt"):
        with open(os.path.join(folder_path, filename), "r") as file:
            lines = file.readlines()
            for i in range(len(lines)):
                if "Product Name" in lines[i]:
                    PD = str(lines[i+1].strip())
                if "Primary Battery Serial Number" not in lines[i]:
                    if "Serial Number" in lines[i]:
                        SN = str(lines[i+1].strip())
                if "System BIOS Version" in lines[i]:
                    BIOS = str(lines[i+1].strip())
                    BIOS_Ver = BIOS[0:17]   # number only = [8:17]
                if "ME Firmware Version" in lines[i]:
                    ME = str(lines[i+1].strip())
                    ME_Ver = ME[0:11]
                if "Feature Byte" in lines[i]:
                    FB = str(lines[i+1].strip())
                    tb1.add_row([PD, SN, BIOS_Ver, ME_Ver, FB])
print(tb1)


def Save_to_TXT():
    with open('Summary.txt', 'w', encoding="utf-8") as f:
        f.write(str(tb1))
        
def Save_to_CSV():
    with open('Summary.csv', 'w', newline='', encoding="utf-8") as w:
        w.write(str(tb1.get_csv_string()))


# CSV 轉 EXCEL
if len(os.listdir(folder_path)) != 0:
    Save_to_CSV()
    
    with open("./Summary.csv") as f:
        raw_data = csv.reader(f)      # 讀取 CSV 檔案
        data = list(raw_data)         # 轉換成串列
    
    wb = Workbook()                   # 建立空白的 Excel 活頁簿物件
    ws = wb.worksheets[0]             # 取得第一個工作表
    ws.title = "BCU_info"             # 命名工作表
    # sheet = wb.create_sheet("TEST") # [選用] 建立一個新的分頁(sheet)並命名
    for i in data:
        ws.append(i)                  # 逐筆添加到工作表內最後一列
 
 
    # 自動對齊欄寬 (可用 但不完美)
    for column_cells in ws.columns:
        new_column_letter = get_column_letter(column_cells[0].column)
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length * 1.21
        

   # 添加邊框
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    for row in ws.rows:
        for cell in row:
            cell.border = thin_border


  # 將第一個row 設置為黑底白字
    fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    font = Font(color='FFFFFF')
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value is not None:
                cell.fill = fill
                cell.font = font


    wb.save('Summary.xlsx')
    os.remove("./Summary.csv")
    
input()